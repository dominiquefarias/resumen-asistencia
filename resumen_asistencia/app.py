import os
import re
import math
import traceback
from flask import Flask, render_template, request, jsonify
import xlrd
import openpyxl

app = Flask(__name__)

def convert_colon_to_comma(val):
    """
    Simplemente reemplaza ':' por ',' manteniendo el texto tal cual (ej. "22:30" -> "22,30").
    """
    if val is None or str(val).strip() == "":
        return "0"

    if isinstance(val, float):
        total_minutes = int(round(val * 24 * 60))
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h},{m:02d}"

    val_str = str(val).strip()
    if ":" in val_str:
        return val_str.replace(":", ",")
    return val_str.replace(".", ",")


def convert_time_to_decimal_comma(val):
    """
    Convierte el tiempo sexagesimal a su equivalente matemático decimal (ej. "22:30" -> "22,5").
    """
    if val is None or str(val).strip() == "":
        return "0"

    if isinstance(val, float):
        total_minutes = val * 24 * 60
        h = int(total_minutes // 60)
        m = total_minutes % 60
        dec_val = h + (m / 60.0)
        factor = 10 ** 3
        truncated = math.trunc(dec_val * factor) / factor
        s = f"{truncated:.3f}".rstrip('0').rstrip('.')
        return s.replace('.', ',') if s else "0"

    val_str = str(val).strip()
    if ":" in val_str:
        parts = val_str.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1])
            dec_val = h + (m / 60.0)
            factor = 10 ** 3
            truncated = math.trunc(dec_val * factor) / factor
            s = f"{truncated:.3f}".rstrip('0').rstrip('.')
            return s.replace('.', ',') if s else "0"
        except ValueError:
            return val_str.replace('.', ',')

    return val_str.replace('.', ',')


def create_empty_worker():
    return {
        'rut': "Sin RUT",
        'nombre': "Sin Nombre",
        'hrs_domingo': "0",
        'hrs_domingo_dec': "0",
        'hrs_extra_domingo': "0",
        'hrs_extra_domingo_dec': "0",
        'horas_festivas': "0",
        'horas_festivas_dec': "0",
        'horas_extras': "0",
        'horas_extras_dec': "0",
        'hrs_descuento': "0",
        'hrs_descuento_dec': "0",
        'ausencias': [],
        'licencias': [],
        'has_summary': False
    }


def process_raw_sheet_matrix(matrix):
    workers = []
    current_worker = None

    for r_idx, row in enumerate(matrix):
        row_str = " ".join([str(cell or "").strip() for cell in row]).upper()

        is_ac_no = any(k in row_str for k in ["AC-NO", "AC - NO", "AC_NO"])
        is_rut = "RUT :" in row_str or "RUT:" in row_str
        is_empleado = "EMPLEADO :" in row_str or "EMPLEADO:" in row_str

        if is_ac_no or (current_worker and current_worker['has_summary'] and (is_rut or is_empleado)):
            if current_worker and (current_worker['nombre'] != "Sin Nombre" or current_worker['rut'] != "Sin RUT"):
                workers.append(finalize_worker(current_worker))
            current_worker = create_empty_worker()

        if not current_worker:
            if is_rut or is_empleado or is_ac_no:
                current_worker = create_empty_worker()
            else:
                continue

        # Extraer RUT y Empleado
        for c_idx, cell in enumerate(row):
            cell_str = str(cell or "").strip()
            cell_upper = cell_str.upper()

            if "RUT" in cell_upper and current_worker['rut'] == "Sin RUT":
                if ":" in cell_str:
                    parts = cell_str.split(":", 1)
                    val = parts[1].strip()
                    if val:
                        current_worker['rut'] = val
                if current_worker['rut'] == "Sin RUT" and c_idx + 1 < len(row):
                    val_right = str(row[c_idx + 1] or "").replace(":", "").strip()
                    if val_right:
                        current_worker['rut'] = val_right

            if "EMPLEADO" in cell_upper and current_worker['nombre'] == "Sin Nombre":
                if ":" in cell_str:
                    parts = cell_str.split(":", 1)
                    val = parts[1].strip()
                    if val:
                        current_worker['nombre'] = val
                if current_worker['nombre'] == "Sin Nombre" and c_idx + 1 < len(row):
                    val_right = str(row[c_idx + 1] or "").replace(":", "").strip()
                    if val_right:
                        current_worker['nombre'] = val_right

        # Capturar Licencias y Faltas por día
        for c_idx, cell in enumerate(row):
            val = str(cell or "").strip()
            val_lower = val.lower()

            if any(val_lower.startswith(day) for day in ['lun', 'mar', 'mie', 'mié', 'jue', 'vie', 'sab', 'sáb', 'dom']):
                match = re.search(r'\b(\d{1,2})\b', val)
                if match:
                    day_num = match.group(1)

                    if "LIC" in row_str and "MEDICA" in row_str:
                        if day_num not in current_worker['licencias']:
                            current_worker['licencias'].append(day_num)
                    elif "FALTA" in row_str or "INASISTENCIA" in row_str:
                        if day_num not in current_worker['ausencias']:
                            current_worker['ausencias'].append(day_num)

        # Capturar valores generando tanto el formato con coma como el cálculo decimal
        for c_idx, cell in enumerate(row):
            lbl = str(cell or "").strip().upper()

            def get_right_values():
                for k in range(c_idx + 1, len(row)):
                    v = row[k]
                    if v is not None and str(v).strip() != "":
                        c_val = convert_colon_to_comma(v)
                        d_val = convert_time_to_decimal_comma(v)
                        return c_val, d_val
                return "0", "0"

            if "HORAS DOMINGO" in lbl:
                current_worker['hrs_domingo'], current_worker['hrs_domingo_dec'] = get_right_values()
                current_worker['has_summary'] = True
            elif "HRS. EXTRA DOMINGO" in lbl or "EXTRA DOMINGO" in lbl:
                current_worker['hrs_extra_domingo'], current_worker['hrs_extra_domingo_dec'] = get_right_values()
                current_worker['has_summary'] = True
            elif "HORAS FESTIVAS" in lbl or "HRS. FESTIVAS" in lbl:
                current_worker['horas_festivas'], current_worker['horas_festivas_dec'] = get_right_values()
                current_worker['has_summary'] = True
            elif "HORAS EXTRAS" in lbl or "HRS. EXTRAS" in lbl:
                current_worker['horas_extras'], current_worker['horas_extras_dec'] = get_right_values()
                current_worker['has_summary'] = True
            elif "HRS. DESCUENTO" in lbl or "HORAS DESCUENTO" in lbl:
                current_worker['hrs_descuento'], current_worker['hrs_descuento_dec'] = get_right_values()
                current_worker['has_summary'] = True

    if current_worker and (current_worker['nombre'] != "Sin Nombre" or current_worker['rut'] != "Sin RUT"):
        workers.append(finalize_worker(current_worker))

    return workers


def finalize_worker(w):
    w['ausencias_str'] = " - ".join(sorted(w['ausencias'], key=lambda x: int(x) if x.isdigit() else 0)) if w['ausencias'] else "Sin faltas"
    w['licencias_str'] = " - ".join(sorted(w['licencias'], key=lambda x: int(x) if x.isdigit() else 0)) if w['licencias'] else "Sin licencias"
    return w


def consolidate_workers(workers):
    final_list = []
    for w in workers:
        if w['nombre'] == "Sin Nombre" and w['rut'] == "Sin RUT" and w['hrs_domingo'] == "0" and w['horas_extras'] == "0":
            continue

        existing = None
        for item in final_list:
            same_rut = (w['rut'] != "Sin RUT" and item['rut'] == w['rut'])
            same_nombre = (w['nombre'] != "Sin Nombre" and item['nombre'] == w['nombre'])
            if same_rut or same_nombre:
                existing = item
                break

        if existing:
            if existing['rut'] == "Sin RUT" and w['rut'] != "Sin RUT":
                existing['rut'] = w['rut']
            if existing['nombre'] == "Sin Nombre" and w['nombre'] != "Sin Nombre":
                existing['nombre'] = w['nombre']

            fields = [
                ('hrs_domingo', 'hrs_domingo_dec'),
                ('hrs_extra_domingo', 'hrs_extra_domingo_dec'),
                ('horas_festivas', 'horas_festivas_dec'),
                ('horas_extras', 'horas_extras_dec'),
                ('hrs_descuento', 'hrs_descuento_dec')
            ]
            for k, k_dec in fields:
                if existing[k] == "0" and w[k] != "0":
                    existing[k] = w[k]
                    existing[k_dec] = w[k_dec]

            if w['ausencias_str'] != "Sin faltas":
                existing['ausencias_str'] = w['ausencias_str'] if existing['ausencias_str'] == "Sin faltas" else f"{existing['ausencias_str']} - {w['ausencias_str']}"
            if w['licencias_str'] != "Sin licencias":
                existing['licencias_str'] = w['licencias_str'] if existing['licencias_str'] == "Sin licencias" else f"{existing['licencias_str']} - {w['licencias_str']}"
        else:
            final_list.append(w)

    return final_list


def parse_report_file(file_path):
    sheets_result = []

    if file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path, formatting_info=False)
        for sheet in wb.sheets():
            sheet_data = []
            for r in range(sheet.nrows):
                row_vals = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                sheet_data.append(row_vals)
            workers = process_raw_sheet_matrix(sheet_data)
            consolidated = consolidate_workers(workers)
            if consolidated:
                sheets_result.append({
                    'sheet_name': sheet.name,
                    'workers': consolidated
                })
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []
            for row in ws.iter_rows(values_only=True):
                sheet_data.append(list(row))
            workers = process_raw_sheet_matrix(sheet_data)
            consolidated = consolidate_workers(workers)
            if consolidated:
                sheets_result.append({
                    'sheet_name': sheet_name,
                    'workers': consolidated
                })

    return sheets_result


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No hay archivo'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    temp_path = os.path.join(os.getcwd(), 'temp_' + file.filename)
    try:
        file.save(temp_path)
        data = parse_report_file(temp_path)
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        traceback.print_exc()
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
